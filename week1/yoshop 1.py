#!/usr/bin/env python
# coding: utf-8

# In[16]:


pip install openpyxl


# In[ ]:


# task 1 1. Python program to Create a excel file


# In[17]:


import pandas as pd


# In[18]:


import openpyxl
# Create a new workbook
workbook = openpyxl.Workbook()

# Get the active worksheet
worksheet = workbook.active

# Add data to the worksheet
worksheet["A1"] = "Name"
worksheet["B1"] = "Age"

worksheet["A2"] = "John"
worksheet["B2"] = 25

worksheet["A3"] = "Emily"
worksheet["B3"] = 30

# Save the workbook
workbook.save("example.xlsx")


# In[19]:


#task 2 2. Python program for Import data from an excel file 
# Load the source Excel file using pandas
source_df = pd.read_excel("Yoshops_Order_List.xlsx")

# Load the target Excel file using openpyxl
target_workbook = openpyxl.load_workbook("example.xlsx")
target_worksheet = target_workbook.active

# Convert the source DataFrame to a list of rows
source_rows = source_df.values.tolist()

# Find the next available row in the target worksheet
next_row = target_worksheet.max_row + 1

# Add data to the target worksheet
for row in source_rows:
    name = row[0]
    age = row[1]
    
    target_worksheet.cell(row=next_row, column=1, value=name)
    target_worksheet.cell(row=next_row, column=2, value=age)
    next_row += 1

# Save the target workbook
target_workbook.save("example.xlsx")


# In[20]:


# Load the target Excel file
target_workbook = openpyxl.load_workbook("example.xlsx")
target_worksheet = target_workbook.active

# Get the data from the target worksheet
data = []
for row in target_worksheet.iter_rows(values_only=True):
    data.append(row)

# Print the data
for row in data:
    print(row)


# In[21]:


import openpyxl
from openpyxl.styles import Font, Alignment

# Load the target Excel file
target_workbook = openpyxl.load_workbook("example.xlsx")
target_worksheet = target_workbook.active

# Format the header row
header_font = Font(bold=True)
header_alignment = Alignment(horizontal="center")
header_row = target_worksheet[1]
for cell in header_row:
    cell.font = header_font
    cell.alignment = header_alignment

# Format the data rows
data_font = Font(italic=True)
data_alignment = Alignment(horizontal="left")
for row in target_worksheet.iter_rows(min_row=2):
    for cell in row:
        cell.font = data_font
        cell.alignment = data_alignment

# Save the target workbook
target_workbook.save("example.xlsx")


# In[26]:


import pandas as pd

# Read the saved Excel file
saved_data = pd.read_excel("example.xlsx")

# Display the data
print(saved_data)


# In[27]:


#Task 4
#4. Python program for Prepare Yoshops Survey and Order excel charts Like = Pie Chart and Bar Chart Weekly, 
#Monthly and Yearly Reports.

import pandas as pd

# Read the Excel file
data = pd.read_excel("Yoshops Survey.xlsx")

# Print the contents of the dataset
print(data)


# In[28]:


# Load the dataset
#data = pd.read_csv("your_dataset.csv")

# Display the original column names
print("1. Name")
print(data.columns)

# Rename a specific column
data.rename(columns={'1. Name': 'Name'}, inplace=True)

# Display the updated column names
print("\nUpdated Column Names:")
print(data.columns)

data


# In[36]:



import pandas as pd
import matplotlib.pyplot as plt

# Load Yoshops Survey dataset
survey_data = data

# Load Yoshops Order dataset
order_data = saved_data

# Combine the datasets based on a common column (e.g., "Name")
combined_data = pd.merge(survey_data, order_data, on="Name")

# Prepare Weekly Survey Report (Pie Chart)
weekly_survey_data = combined_data["6. Laptop and Mobile which Price range you like most"].value_counts()
fig, ax = plt.subplots()
weekly_survey_data.plot(kind='pie', autopct='%1.1f%%', ax=ax)
ax.set_title('Weekly Survey Report')


# In[42]:


combined_data = pd.merge(survey_data, order_data, left_on="Name", right_on="Name")


# In[44]:


combined_data


# In[43]:


survey_data.rename(columns={'Survey_Submitted_Time': 'Submitted Time'}, inplace=True)
combined_data = pd.merge(survey_data, order_data, left_on="Name", right_on="Name")


# In[46]:


fig, ax = plt.subplots()
monthly_survey_data.iloc[:, 0].plot(kind='pie', autopct='%1.1f%%', ax=ax)
ax.set_title('Monthly Survey Report')


# In[47]:


fig, ax = plt.subplots(nrows=1, ncols=len(monthly_survey_data.columns))
for i, column in enumerate(monthly_survey_data.columns):
    monthly_survey_data[column].plot(kind='pie', autopct='%1.1f%%', ax=ax[i])
    ax[i].set_title(column)
plt.suptitle('Monthly Survey Report')
plt.tight_layout()


# In[48]:


fig, ax = plt.subplots()
monthly_survey_data.plot(kind='bar', ax=ax)
ax.set_title('Monthly Survey Report')
ax.set_xlabel('Month')
ax.set_ylabel('Count')
plt.xticks(rotation=45)
plt.tight_layout()


# In[49]:


fig, ax = plt.subplots()
yearly_survey_data = combined_data.groupby(combined_data.index.year)['6. Laptop and Mobile which Price range you like most'].value_counts().unstack()
yearly_survey_data.plot(kind='bar', ax=ax)
ax.set_title('Yearly Survey Report')
ax.set_xlabel('Year')
ax.set_ylabel('Count')
plt.xticks(rotation=45)
plt.tight_layout()


# In[53]:


pip install python-docx


# In[70]:


#task 5
import json
import pandas as pd
import zipfile

def extract_mobile_numbers_from_json(file_path):
    # Open the ZIP file
    with zipfile.ZipFile(file_path, 'r') as zip_file:
        # Find the JSON file within the ZIP archive (assuming there's only one JSON file)
        json_file_name = next(file for file in zip_file.namelist() if file.endswith('.json'))

        # Read the JSON file from the ZIP archive
        with zip_file.open(json_file_name) as file:
            json_data = json.load(file)

    # Extract mobile numbers from the JSON data
    mobile_numbers = []
    for item in json_data:
        if 'mobile_number' in item:
            mobile_numbers.append(item['mobile_number'])

    # Create a DataFrame from the mobile numbers
    df = pd.DataFrame({'Mobile Number': mobile_numbers})

    # Save the DataFrame to an Excel file
    output_file = file_path.replace('.zip', '_mobile_numbers.xlsx')
    df.to_excel(output_file, index=False)
    print(f"Mobile numbers extracted and saved to {output_file}")

# Example usage
file_path = "C:\\Users\\athira shaji\\Desktop\\yoshop\\contact data.zip"
extract_mobile_numbers_from_json(file_path)


# In[73]:


print(data.columns)


# In[77]:


# task 6
import pandas as pd

data = pd.read_excel("Yoshops Survey.xlsx")

# Drop unnecessary columns
columns_to_drop = ['S.NO', '3. Location , City Name', '7. What is your Favourite food biryani essay?',
                   '9. Which Financial products you like most', '10. Are you interested in Yoshops Finance  Product.',
                   '11. Which Unpaid Training Internship are you interested for 6 Months ?',
                   '12. How much you pay training fees per month for online training with internship?',
                   '15. Do you want to get a free Gift and Refferal Amount Rs.100 by reffer friends to Training Internship and Online Tituation',
                   ]

# Remove leading and trailing spaces from column names
columns_to_drop = [col.strip() for col in columns_to_drop]

data_cleaned = data.drop(columns=columns_to_drop)

# Remove rows with missing values
data_cleaned = data_cleaned.dropna()

# Convert 'Submitted Time' column to datetime
data_cleaned['Submitted Time'] = pd.to_datetime(data_cleaned['Submitted Time'], dayfirst=True)

# Save the cleaned data to a new CSV file
data_cleaned.to_csv('cleaned_data.csv', index=False)
print("Data cleaning process completed and cleaned data saved to 'cleaned_data.csv'")


# In[ ]:




